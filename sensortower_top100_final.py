"""
SensorTower Top-100 Photo & Video (US App Store) — выгрузка в Excel
=======================================================================
Запуск:    python3 sensortower_top100_final.py
Зависимости: pip3 install requests pandas openpyxl
"""

import time
import requests
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────
#  НАСТРОЙКИ — обновить если сессия истекла
# ─────────────────────────────────────────────────────────────
COOKIE = (
    "AMP_6edb64137a=JTdCJTIyZGV2aWNlSWQlMjIlM0ElMjI5MjRjM2IzMS03NzI0LTRhNTItYWE0YS03ODY5Mzc3MzQ5ZmQlMjIl"
    "MkMlMjJ1c2VySWQlMjIlM0ElMjJnb21vem92Lml2YW4uOTglNDBnbWFpbC5jb20lMjIlMkMlMjJzZXNzaW9uSWQlMjIlM0ExNzc3"
    "NDc5MjA3MTMyJTJDJTIyb3B0T3V0JTIyJTNBZmFsc2UlMkMlMjJsYXN0RXZlbnRUaW1lJTIyJTNBMTc3NzQ3OTM1MDM3NyUyQyUy"
    "Mmxhc3RFdmVudElkJTIyJTNBNDIlMkMlMjJwYWdlQ291bnRlciUyMiUzQTAlN0Q=; "
    "_ga=GA1.1.1774347708.1777470489; "
    "_ga_FDNER2EVFL=GS2.1.s1777479319$o2$g1$t1777479349$j30$l0$h0; "
    "osano_consentmanager=GWvFrq8rHN7ZK5YtBP7YBl5WR88ZCpSZvbnYHNeTUIb0Xyk7v0o2YhwaINkFYtnQI4airN2J4yhkkjoecN60"
    "GoK2ltq0k0eQRg2IeiPyLN0y0Gu7p-kVY04WHAajrwz8L2EY9xWPAxCNHbRaoqRmY3i2zO3UqhPTTc7chz6su6hXLzStqz627kIXn"
    "pLJYv-nEvRkTF72oC0vGJlbaboSg1BsWO2kskvsi3OhKTq1W0434Fmw5idpn5thqByGNmy0QDymcU5qOAt_WuXn0Qxiv5X7oehPIm7"
    "XXl4Zz3PZ7SYtdAdSvgo4M-x7MSPh6cd2LBW8-XDmPN0=; "
    "osano_consentmanager_uuid=729a8888-30a8-4b55-bbf0-d2c45e247c6b; "
    "sensor_tower_session=7073f14cecc2d0872c16ad0d3b471c18; "
    "locale=en"
)
CSRF_TOKEN  = "fl-dYBwwIUpuNJjiX6knubwEeODS4-dZSZcVQSO2NVTB52dxNu8uS6lOBQB7pZrURG_2zUwRBEHb0Ps_mwP1WQ"

CATEGORY    = "6008"
COUNTRY     = "US"
DATE        = "2026-04-20"
DEVICE      = "iphone"
CHART_TYPE  = "free"
OUTPUT_FILE = "sensortower_top100.xlsx"
# ─────────────────────────────────────────────────────────────

BASE_URL = "https://app.sensortower.com/api/ios/category_rankings"
HEADERS  = {
    "Accept":       "application/json",
    "Content-Type": "application/json",
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.6 Safari/605.1.15"
    ),
    "Referer":      "https://app.sensortower.com/top-charts",
    "Cookie":       COOKIE,
    "X-CSRF-Token": CSRF_TOKEN,
}

# Числовые форматы Excel — значения хранятся как числа (сортируются корректно),
# отображаются как "13M" / "$193M" / "$7.50"
FMT_DL  = '[>=1000000]0.0,,"M";[>=1000]0.0,"K";0'
FMT_REV = '[>=1000000]"$"0.0,,"M";[>=1000]"$"0.0,"K";"$"0'
FMT_RPD = '"$"0.00'


# ── Fetch ──────────────────────────────────────────────────────
def fetch_page(offset: int, limit: int = 25) -> list:
    params = dict(offset=offset, limit=limit, category=CATEGORY,
                  country=COUNTRY, date=DATE, device=DEVICE)
    r = requests.get(BASE_URL, headers=HEADERS, params=params, timeout=15)
    r.raise_for_status()
    return r.json()["data"].get(CHART_TYPE, [])

def fetch_all(total: int = 100, page_size: int = 25) -> list:
    apps, offsets = [], range(0, total, page_size)
    for i, offset in enumerate(offsets, 1):
        print(f"  Страница {i}/{len(offsets)} (offset={offset})…")
        apps.extend(fetch_page(offset, page_size))
        if offset + page_size < total:
            time.sleep(0.3)
    return apps


# ── Parse ──────────────────────────────────────────────────────
def parse(raw: list) -> pd.DataFrame:
    rows = []
    for app in raw:
        dl  = app["humanized_worldwide_last_month_downloads"]["downloads"]
        rev = app["humanized_worldwide_last_month_revenue"]["revenue"]
        rpd = round(rev / dl, 2) if dl else 0
        rows.append({
            "#":          app["rank"],
            "Приложение": app["name"],
            "Издатель":   app["publisher_name"],
            "_dl":        dl,
            "_rev":       rev,
            "_rpd":       rpd,
        })
    df = pd.DataFrame(rows).sort_values("#").reset_index(drop=True)

    # Метрика 1: топ-10 по $ / установка (зелёный)
    df["_mon_rank"] = df["_rpd"].rank(ascending=False, method="min").astype(int)

    # Метрика 2: нижние 33 строки по загрузкам → строго топ-10 по выручке среди них (синий)
    bottom33_idx  = df.nsmallest(33, "_dl").index
    top10_idx     = df.loc[bottom33_idx].nlargest(10, "_rev", keep="first").index
    df["_blue"]   = df.index.isin(top10_idx)
    assert df["_blue"].sum() == 10, f"Ошибка: синих строк {df['_blue'].sum()}, ожидалось 10"

    return df


# ── Excel ──────────────────────────────────────────────────────
C_NAVY  = "1B3A6B"
C_WHITE = "FFFFFF"
C_LGRAY = "F7F8FA"
C_GREEN = "D6EFD8"
C_BLUE  = "D6E8FA"
C_BOTH  = "C8E6C9"
C_BLACK = "1A1A1A"
_thin   = Side(style="thin",   color="D0D5DD")
_med    = Side(style="medium", color=C_NAVY)

def brd(ci: int, n: int, last_row: bool = False) -> Border:
    return Border(
        left   = _med if ci == 1 else _thin,
        right  = _med if ci == n else _thin,
        top    = _thin,
        bottom = _med if last_row else _thin,
    )

def save_excel(df: pd.DataFrame, path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Top-100 Photo & Video"
    N = 6
    total_rows = len(df)

    # Row 1: title
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value     = f"Top-100 Photo & Video Apps  |  US App Store  |  {DATE}  |  Source: Sensor Tower"
    c.font      = Font(name="Arial", bold=True, size=11, color=C_WHITE)
    c.fill      = PatternFill("solid", fgColor=C_NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 28

    # Row 2: legend
    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value     = "  🟢  Топ по $ / установка   |   🔵  Топ по выручке среди приложений с минимальными загрузками"
    c.font      = Font(name="Arial", italic=True, size=9, color="1A3A6B")
    c.fill      = PatternFill("solid", fgColor="EAF2FB")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 18

    # Row 3: headers
    HEADERS = ["#", "Приложение", "Издатель",
               "Загрузки\n(мес.)", "Выручка\n(мес.)", "$ / установка"]
    for ci, h in enumerate(HEADERS, 1):
        c = ws.cell(row=3, column=ci, value=h)
        c.font      = Font(name="Arial", bold=True, size=10, color=C_WHITE)
        c.fill      = PatternFill("solid", fgColor=C_NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = brd(ci, N)
    ws.row_dimensions[3].height = 34

    # Data rows
    for ri, row in df.iterrows():
        er      = ri + 4
        is_last = (ri == total_rows - 1)
        green   = row["_mon_rank"] <= 10
        blue    = row["_blue"]

        if green and blue: bg = C_BOTH
        elif green:        bg = C_GREEN
        elif blue:         bg = C_BLUE
        else:              bg = C_LGRAY if ri % 2 == 0 else C_WHITE
        fill = PatternFill("solid", fgColor=bg)

        # # (ранг)
        c = ws.cell(row=er, column=1, value=int(row["#"]))
        c.fill = fill; c.border = brd(1, N, is_last)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.font = Font(name="Arial", size=10, bold=True, color=C_NAVY)

        # Приложение
        c = ws.cell(row=er, column=2, value=row["Приложение"])
        c.fill = fill; c.border = brd(2, N, is_last)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c.font = Font(name="Arial", size=10, bold=True, color=C_BLACK)

        # Издатель
        c = ws.cell(row=er, column=3, value=row["Издатель"])
        c.fill = fill; c.border = brd(3, N, is_last)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c.font = Font(name="Arial", size=10, color=C_BLACK)

        # Загрузки — число, отображается как "13M" (сортируется как число)
        c = ws.cell(row=er, column=4, value=int(row["_dl"]))
        c.number_format = FMT_DL
        c.fill = fill; c.border = brd(4, N, is_last)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.font = Font(name="Arial", size=10, color=C_BLACK)

        # Выручка — число, отображается как "$193M" (сортируется как число)
        c = ws.cell(row=er, column=5, value=int(row["_rev"]))
        c.number_format = FMT_REV
        c.fill = fill; c.border = brd(5, N, is_last)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.font = Font(name="Arial", size=10, color=C_BLACK)

        # $ / установка
        c = ws.cell(row=er, column=6, value=float(row["_rpd"]))
        c.number_format = FMT_RPD
        c.fill = fill; c.border = brd(6, N, is_last)
        c.alignment = Alignment(horizontal="center", vertical="center")
        if green:
            c.font = Font(name="Arial", size=10, bold=True, color="1A6B2E")
        elif blue:
            c.font = Font(name="Arial", size=10, bold=True, color="1A4A8B")
        else:
            c.font = Font(name="Arial", size=10, color=C_BLACK)

        ws.row_dimensions[er].height = 18

    # Column widths
    for ci, w in enumerate([5, 34, 26, 13, 13, 15], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:F{3 + len(df)}"

    # Методология
    ws2 = wb.create_sheet("Методология")
    for r in [
        ("Поле / Выделение",     "Описание"),
        ("Загрузки (мес.)",      "Суммарное число установок приложения за последний месяц (worldwide)"),
        ("Выручка (мес.)",       "Выручка от In-App Purchases + платных загрузок за месяц (worldwide, gross, до комиссии Apple)"),
        ("$ / установка",        "Выручка ÷ Загрузки — средний доход с одной установки"),
        ("🟢 Зелёная подсветка", "Приложения с наибольшим показателем $ / установка среди всех 100"),
        ("🔵 Синяя подсветка",   "Приложения с наибольшей выручкой среди тех, у кого минимальные загрузки (нижняя треть по загрузкам)"),
    ]:
        ws2.append(r)
    ws2["A1"].font = Font(name="Arial", bold=True, size=10)
    ws2["B1"].font = Font(name="Arial", bold=True, size=10)
    for row in ws2.iter_rows(min_row=2):
        for cell in row:
            cell.font      = Font(name="Arial", size=10)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 85
    for i in range(1, 8):
        ws2.row_dimensions[i].height = 32

    wb.save(path)
    print(f"\n✅  Сохранено: {path}")


# ── Main ───────────────────────────────────────────────────────
def main():
    print("═" * 55)
    print("  SensorTower Top-100 Photo & Video — US App Store")
    print(f"  Дата: {DATE}  |  Чарт: {CHART_TYPE.upper()}")
    print("═" * 55)

    print("\n📡  Загружаем данные…")
    raw = fetch_all(total=100, page_size=25)
    print(f"   Получено: {len(raw)} записей")

    df = parse(raw)

    print("\n🟢  Топ-10 по $ / установка:")
    for _, r in df[df["_mon_rank"] <= 10].sort_values("_mon_rank").iterrows():
        print(f"  #{int(r['_mon_rank']):2d}  {r['Приложение']:<35s}  "
              f"dl={r['_dl']:>10,}  rev=${r['_rev']:>10,}  ${r['_rpd']:.2f}/install")

    print("\n🔵  Топ-10 по выручке среди приложений с мин. загрузками:")
    for _, r in df[df["_blue"]].sort_values("_rev", ascending=False).iterrows():
        print(f"  #{int(r['#']):3d}  {r['Приложение']:<35s}  "
              f"dl={r['_dl']:>10,}  rev=${r['_rev']:>10,}")

    print(f"\n💾  Сохраняем {OUTPUT_FILE}…")
    save_excel(df, OUTPUT_FILE)


if __name__ == "__main__":
    main()
